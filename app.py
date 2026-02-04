from flask import Flask, request, redirect, render_template_string, send_file
from openpyxl import Workbook, load_workbook
import pandas as pd
import os

app = Flask(__name__)

FILE_NAME = "reject_data.xlsx"

HEADERS = [
    "Date", "Process", "Problem Type",
    "Reject Quantity", "Main Customer", "Workcenter"
]

PROCESSES = [
    "UV HOCK","VARNISH","LAMINATING","WP GERMAN","WP ESATEC",
    "WP ZHENGMAO 1 C","WP ZHENGMAO 2 C",
    "FG MATTEL","FG MASTER","FG BICHENG","FG BOBST",
    "JHOOK 5 C","JHOOK 1",
    "LINE 1","LINE 2","LINE 3","LINE 4",
    *[f"PH {i}" for i in range(1,23)],
    *[f"HS {i}" for i in range(1,7)],
    "SABLON SEMI AUTO","LINE BORONGAN",
    "FG ROLAM C",
    "VACUUM 1","VACUUM 2","VACUUM 3","VACUUM 4"
]

# BUAT FILE EXCEL JIKA BELUM ADA
if not os.path.exists(FILE_NAME):
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    wb.save(FILE_NAME)

HTML = """
<h2>SERVER INPUT REJECT</h2>

<h3>Input Manual</h3>
<form method="post">
Date <input type="date" name="date" required>
Process
<select name="process">
{% for p in processes %}
<option>{{p}}</option>
{% endfor %}
</select>
Problem <input name="problem" required>
Qty <input type="number" name="qty" required>
Customer <input name="customer" required>
Workcenter <input name="workcenter" required>
<button>Add</button>
</form>

<hr>

<h3>Upload Excel</h3>
<form action="/upload" method="post" enctype="multipart/form-data">
<input type="file" name="file" accept=".xlsx" required>
<button>Upload</button>
</form>

<hr>

<a href="/download">Download Excel</a>

<form action="/delete_all" method="post" style="margin-top:10px">
<button>DELETE ALL DATA</button>
</form>

<hr>

<table border="1" cellpadding="5">
<tr>
{% for h in headers %}
<th>{{h}}</th>
{% endfor %}
<th>Action</th>
</tr>

{% for row in data %}
<tr>
{% for c in row %}
<td>{{c}}</td>
{% endfor %}
<td>
<form action="/delete/{{ loop.index0 }}" method="post">
<button>Delete</button>
</form>
</td>
</tr>
{% endfor %}
</table>
"""

@app.route("/", methods=["GET","POST"])
def index():
    wb = load_workbook(FILE_NAME)
    ws = wb.active

    if request.method == "POST":
        ws.append([
            request.form["date"],
            request.form["process"],
            request.form["problem"],
            request.form["qty"],
            request.form["customer"],
            request.form["workcenter"]
        ])
        wb.save(FILE_NAME)
        return redirect("/")

    data = list(ws.iter_rows(values_only=True))[1:]

    return render_template_string(
        HTML,
        data=data,
        headers=HEADERS,
        processes=PROCESSES
    )

@app.route("/upload", methods=["POST"])
def upload():
    file = request.files["file"]
    df = pd.read_excel(file)

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    for row in df.values.tolist():
        ws.append(row)

    wb.save(FILE_NAME)
    return redirect("/")

@app.route("/delete/<int:row_id>", methods=["POST"])
def delete_row(row_id):
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    ws.delete_rows(row_id + 2)
    wb.save(FILE_NAME)
    return redirect("/")

@app.route("/delete_all", methods=["POST"])
def delete_all():
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    ws.delete_rows(2, ws.max_row)
    wb.save(FILE_NAME)
    return redirect("/")

@app.route("/download")
def download():
    return send_file(FILE_NAME, as_attachment=True)

if __name__ == "__main__":
    app.run(debug=False)
